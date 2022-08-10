import folium as fl
from openpyxl import load_workbook

def tuple_to_list(x):
    return [x[y].value for y in range(len(x))]

# loading excel file into vsc
wb = load_workbook('cpt_restaurants.xlsx')
ws = wb.worksheets[0]
nme = ws['A']
lat = ws['B']
long = ws['C']
addrs = ws['D']
img = ws['E']
urls = ws['F']
kor = ws['G']

# pulling data from columns into lists 
name_list = tuple_to_list(nme)
lat_list = tuple_to_list(lat)
long_list = tuple_to_list(long)
addrs_list = tuple_to_list(addrs)
img_list = tuple_to_list(img)
url_list = tuple_to_list(urls)
kor_list = tuple_to_list(kor)

# instantiating map object 
map = fl.Map(location=[-33.92714772961316, 18.41353385511342], zoom_start=13, tiles="Stamen Terrain")

# creating feature group 
fg = fl.FeatureGroup(name="myMap")

# placing markers at favourite restaurants using data from imported xlsx file
for nm, ad, im, ur, lt, ln, kr in zip(name_list, addrs_list, img_list, url_list, lat_list, long_list, kor_list):
   
    # creating html for the map markers
    html = f""" 
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.2.0-beta1/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-0evHe/X+R7YkIZDRvuzKMRqM+OrBnVFBL6DOitfPri4tjfHxaWutUpFmBp4vmVor" crossorigin="anonymous">
            </head>
            <div style="display:inline-block">
            <center>
                <h1 style="color:lightskyblue;"> {nm} </h1>
            </center>
            <center>
                <h3 style="color:lightgrey"> {kr} </h3>
            </center>
            <center>                
                <p> Address: </p>
                <p><em> {ad} </em></p>
            </center>
            <center>
                <img src="{im}" style="height:auto; max-width:100%; margin:20px;" alt="Image of {nm}"></img>
            </center>
            <center>
                <a href="{ur}" type="button" target="_blank" rel="noopener" class="btn btn-light btn-block">Visit Page</a>            
            </center>
            </div>
            </html>
            """
    iframe = fl.IFrame(html, width = 250, height = 400)
    popup = fl.Popup(iframe)
    fg.add_child(fl.Marker(location = [lt, ln], popup = (popup) ))

map.add_child(fg)

map.save("index.html")
